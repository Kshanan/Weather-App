import tkinter as tk
from tkinter import END,messagebox
from tkinter.font import BOLD
import requests
from PIL import Image,ImageTk #pip install pillow    
import datetime 
import openpyxl    #pip install openpyxl
from openpyxl import Workbook   
import pathlib2    #pip install pathlib2
import socket
import speech_recognition as sr



#Creating an Empty interface with title as Weather App 1.0
root=tk.Tk()   
root.title("Weather App 1.0")
root.geometry("500x400")

#Looking for Existing Excel sheet and Editing Defaults
file= pathlib2.Path("Weather_data.xlsx")   
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]="Date"
    sheet["B1"]="Time"
    sheet["C1"]="City"
    sheet["D1"]="Temperature in celsius"
    file.save("Weather_data.xlsx")

#Storing data in the Excel sheet
def store(city):
    weather_key='b171ec8256ff1b1ed07b680ada20ddbc'
    url='https://api.openweathermap.org/data/2.5/weather'
    params={'APPID':weather_key,'q':city,'units':'metric'}
    response=requests.get(url,params)
    weather=response.json()        
    tz=datetime.timezone(datetime.timedelta(seconds=int(weather['timezone'])))
    tx=datetime.datetime.now(tz=tz).strftime("%d/%m/%Y")
    date=tx
    tu=datetime.datetime.now(tz=tz).strftime("%I:%M %p")
    time=tu
    city=weather['name']
    temp=weather['main']['temp']
    file=openpyxl.load_workbook("Weather_data.xlsx")
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=date)
    sheet.cell(column=2,row=sheet.max_row,value=time)
    sheet.cell(column=3,row=sheet.max_row,value=city)
    sheet.cell(column=4,row=sheet.max_row,value=temp)
    file.save("Weather_data.xlsx")
    messagebox.showinfo('SAVE',"Data Saved")
    result['text']='Check Data in Excel'
     

#interpreting API received data
#All the result stored in the final_str variable (returns these value when called)

def format_response(weather):   
    try:      
        city=weather['name']
        condition=weather['weather'][0]['description']
        tz=datetime.timezone(datetime.timedelta(seconds=int(weather['timezone'])))
        tx=datetime.datetime.now(tz=tz).strftime("%d/%m/%Y, %I:%M %p  ")
        time=tx
        temp=weather['main']['temp']
        feels_like=weather['main']['feels_like']
        humidity=weather['main']['humidity']
        final_str=('City : %s\nCondition : %s\n Date & Time : %s\nTemprature : %s °c\nFeels like : %s °c\nHumidity : %s \n'%(city,condition,time,temp,feels_like,humidity))
    except:
        final_str =messagebox.showwarning('Error 404',"City not found")
    return final_str      

# Api call url : api.openweathermap.org/data/2.5/weather?q={city name}&appid={API key}
#Api Key : 2f46ace604098d5610d752ab3d07a893
#Api Key2 : f33fedc981179a0c9c56df0023c70085

#Testing Internet Connectivity 

def test_con():                                                        
    try:
        socket.create_connection(('google.com',80))
        return get_weather(txt_box.get())
    except OSError:
        messagebox.showwarning('Warning',"No Internet Connection")
        return result


#Fetching API data with url and API key    
def get_weather(city):           
    weather_key='b171ec8256ff1b1ed07b680ada20ddbc'
    url='https://api.openweathermap.org/data/2.5/weather'
    params={'APPID':weather_key,'q':city,'units':'metric'}
    response=requests.get(url,params)
    weather=response.json()
     
    result['text']=format_response(weather)    #Result received after interpreting the API data
                                              
    icon_name=weather['weather'][0]['icon']
    open_image(icon_name)

#To diaplay the icon images as per result from API data  
def open_image(icon):       
    size=int(frame_two.winfo_height()*0.40)
    img=ImageTk.PhotoImage(Image.open('./img/'+icon+'.png').resize((size,size)))
    weather_icon.delete('all')
    weather_icon.create_image(0,0,anchor='nw',image=img)
    weather_icon.image=img

#MoreInfo 

def info():
    inf=messagebox.showinfo("More Info","This app has been developed only as a prototype for Academic Projects ,The project belongs to Vth sem BCA(Co-Ed).                                                                              Project submitted by :Kshanan Muttin SG19BCA063   | Manjunath SG19BCA077  |   Mohammed Azeem SG19BCA080   |   Mehboob Paatel SG19BCA072                                                Guided by : Sharanu Sagare Sir")
    return inf

#Setting up Voice Recognition

r = sr.Recognizer()

def get_voice():
    
    try:
        with sr.Microphone() as source:
            audio = r.listen(source)
            text=r.recognize_google(audio)
            txt_box.delete(0,END)
            txt_box.insert(0,text)
            test_con()
    except:
        result['text']='sorry could not recognise what you said'
            


#Fetching background image for the interface
#Setting the fetching image as Background for the interface
img=Image.open('./dd.jpg')     
img=img.resize((500,400),Image.ANTIALIAS)
img_photo=ImageTk.PhotoImage(img)
bg_lbl=tk.Label(root,image=img_photo)  
bg_lbl.place(x=0,y=0,width=500,height=400)

#Adding a heading line 
heading_title=tk.Label(bg_lbl,text='Check Weather of over 200,000 cities around the globe!',fg='black',bg='white',font=('times new roman',14,'bold'))
heading_title.place(x=22,y=14)

#creating a empty frame
frame_one=tk.Frame(bg_lbl,highlightbackground='black',highlightthickness=2)
frame_one.place(x=40,y=50,width=310,height=40)

#Adding textbox to frame for search bar
txt_box=tk.Entry(frame_one,font=('times new roman',22),width=21)
txt_box.grid(row=0,column=0,sticky='w')

#Adding search button
img1=Image.open('./see.png')
img1=img1.resize((100,40),Image.ANTIALIAS)
img2=ImageTk.PhotoImage(img1)
btn=tk.Button(root,image=img2,bg='white',borderwidth=0,command=lambda:test_con())
btn.place(x=360,y=52,width=100,height=40)

#Button to save the data in Excel sheet  
img7=Image.open('./sve2.jpg')
img7=img7.resize((60,25),Image.ANTIALIAS)
img_sv=ImageTk.PhotoImage(img7)
btn=tk.Button(root,image=img_sv,bg='white',borderwidth=0,command=lambda:store(txt_box.get()))
btn.place(x=40,y=355,width=60,height=25)

#Creating another frame to display the interpreted result
frame_two=tk.Frame(bg_lbl,highlightbackground='black',highlightthickness=2)
frame_two.place(x=40,y=100,width=400,height=250)

#Displaying the result 
result=tk.Label(frame_two,font=('Times new roman',10,BOLD),bg='white',justify='center',anchor='s')
result.place(relwidth=1,relheight=1)

#Displaying the Weather icon as per the result obtained from the API 
weather_icon=tk.Canvas(result,bg='white',bd=0,highlightthickness=0)
weather_icon.place(relx=0.38,rely=0.12,relwidth=1,relheight=0.5)

#label to display the credits
info=tk.Label(bg_lbl,text='project by BCA(Co-Ed)',fg='cyan',bg='black',font=('Brush Script MT',12))
info.place(x=320,y=355)

#Moreinfo Button
img3=Image.open('./hy.png')
img3=img3.resize((60,25),Image.ANTIALIAS)    
img4=ImageTk.PhotoImage(img3)
btn2=tk.Button(root,image=img4,borderwidth=0,command=info)
btn2.place(x=200,y=355,width=60,height=25) 

#Mic Button
img_mic=Image.open('./mic.png')
img_mic=img_mic.resize((30,30),Image.ANTIALIAS)    
img_m=ImageTk.PhotoImage(img_mic)
btn3=tk.Button(root,image=img_m,borderwidth=0,command=lambda:get_voice())
btn3.place(x=10,y=55,width=30,height=30) 


      

 
root.mainloop()

